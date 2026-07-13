"""
Part 1: openpyxl — build base .xlsx
  - synagogues: removed originalNusach + removed all day-prayer columns
  - new sheet: תפילות_מפורטות (prayer times with anchor/offset support)
  - all other sheets unchanged
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FILL  = PatternFill("solid", start_color="DDEEFF")
EXAMPLE_FILL = PatternFill("solid", start_color="F5F5F5")
INSTR_ROW    = PatternFill("solid", start_color="E8F0FE")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")
SECTION_FILL = PatternFill("solid", start_color="2E6DB4")
TITLE_FILL   = PatternFill("solid", start_color="1B4E8A")
SUB_FILL     = PatternFill("solid", start_color="2E6DB4")

HEADER_FONT  = Font(name="Arial", bold=True, color="1B3A6B", size=10)
EXAMPLE_FONT = Font(name="Arial", italic=True, color="888888", size=9)
TITLE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=13)
SEC_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=10)
INSTR_BOLD   = Font(name="Arial", bold=True, size=10, color="1B3A6B")
INSTR_NORM   = Font(name="Arial", size=10, color="1B1B1B")
INSTR_ITALIC = Font(name="Arial", italic=True, size=9, color="666666")

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(cell):
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def ex(cell):
    cell.font = EXAMPLE_FONT; cell.fill = EXAMPLE_FILL; cell.border = BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def add_sheet(name, columns, example_row):
    ws = wb.create_sheet(name)
    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 18
    for ci, (en, he, _note, w) in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=f"{he}\n{en}")
        hdr(c)
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ci, val in enumerate(example_row, 1):
        c = ws.cell(row=2, column=ci, value=val)
        ex(c)
    ws.freeze_panes = "A2"
    return ws


# ── Instructions ─────────────────────────────────────────────────────────────
ws_i = wb.active; ws_i.title = "הוראות"
ws_i.sheet_view.rightToLeft = True
ws_i.column_dimensions["A"].width = 4
ws_i.column_dimensions["B"].width = 26
ws_i.column_dimensions["C"].width = 52
ws_i.column_dimensions["D"].width = 36

ws_i.merge_cells("A1:D1"); ws_i["A1"] = "מדריך מילוי נתונים — אפליקציית קהילה"
ws_i["A1"].font = TITLE_FONT; ws_i["A1"].fill = TITLE_FILL
ws_i["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_i.row_dimensions[1].height = 34

ws_i.merge_cells("A2:D2"); ws_i["A2"] = "לחצו על כפתור ➕ בכל גיליון להוספת רשומה באמצעות טופס נוח"
ws_i["A2"].font = Font(name="Arial", italic=True, color="FFFFFF", size=10)
ws_i["A2"].fill = SUB_FILL
ws_i["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws_i.row_dimensions[2].height = 20

def sec(row, title):
    ws_i.merge_cells(f"A{row}:D{row}")
    c = ws_i[f"A{row}"]; c.value = "  " + title
    c.font = SEC_FONT; c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws_i.row_dimensions[row].height = 24

def ir(row, label, value, note=""):
    ws_i[f"B{row}"] = label
    ws_i[f"B{row}"].font = INSTR_BOLD; ws_i[f"B{row}"].fill = INSTR_ROW
    ws_i[f"B{row}"].border = BORDER
    ws_i[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws_i[f"C{row}"] = value
    ws_i[f"C{row}"].font = INSTR_NORM; ws_i[f"C{row}"].fill = WHITE_FILL
    ws_i[f"C{row}"].border = BORDER
    ws_i[f"C{row}"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    if note:
        ws_i[f"D{row}"] = note
        ws_i[f"D{row}"].font = INSTR_ITALIC; ws_i[f"D{row}"].fill = WHITE_FILL
        ws_i[f"D{row}"].border = BORDER
        ws_i[f"D{row}"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws_i.row_dimensions[row].height = 20

r = 4
sec(r, "כללי")
r+=1; ir(r,"כפתור ➕","לחצו על הכפתור הכחול ➕ בפינה השמאלית של כל גיליון לפתיחת טופס הזנה")
r+=1; ir(r,"שורה לדוגמה","שורה 2 בכל גיליון מוצגת באפור — ניתן לשנות או למחוק")
r+=1; ir(r,"מזהה (ID)",'מזהה ייחודי: syn-001, rest-001, mik-001, evt-001, shi-001, pry-001',"אל תשתמשו ברווחים")
r+=1; ir(r,"שדות חובה","שדות המסומנים * הם חובה")

r+=1; sec(r, "פורמטים")
r+=1; ir(r,"תאריך","YYYY-MM-DD","דוגמה: 2025-09-14")
r+=1; ir(r,"שעה","HH:MM  (24 שעות)","דוגמה: 18:30")
r+=1; ir(r,"קואורדינטות","מספר עשרוני עם נקודה","דוגמה: 31.9234")
r+=1; ir(r,"שעות פתיחה","HH:MM-HH:MM  או  סגור","דוגמה: 09:00-22:00")

r+=1; sec(r, "ערכי שדות")
r+=1; ir(r,"nusach","ashkenaz / sefard / edot_hamizrach / maroko / other")
r+=1; ir(r,"קטגוריה מסעדה","meat / dairy / pareve / cafe / bakery")
r+=1; ir(r,"קטגוריה אירוע","shiur / community / youth / charity / holiday / announcement / alert")
r+=1; ir(r,"ימים","1=ראשון  2=שני  3=שלישי  4=רביעי  5=חמישי  6=שישי  7=שבת  —  הפרידו בפסיק","דוגמה: 1,2,3,4,5")
r+=1; ir(r,"סוג זמן (תפילות)","weekday = ימות השבוע (ימים 1-6)  |  shabbat = שבת  |  friday_mincha = מנחה ע\"ש")
r+=1; ir(r,"עוגן זמנים","netz / shkia / chatzot / minchaGedola / minchaKetana / plagHamincha","לזמן יחסי כגון 20 דק' לפני שקיעה")
r+=1; ir(r,"קיזוז (offsetMin)","מספר דקות — חיובי = אחרי העוגן  |  שלילי = לפני העוגן","דוגמה: -20 = 20 דק' לפני")

r+=1; sec(r, "גיליונות")
r+=1; ir(r,"בתי_כנסת","מידע בסיסי על בתי כנסת")
r+=1; ir(r,"תפילות_מפורטות","זמני תפילה מפורטים — ניתן להגדיר זמן יחסי לזמנים יהודיים")
r+=1; ir(r,"מסעדות_כשרות","מסעדות ועסקים כשרים + תעודות כשרות")
r+=1; ir(r,"מקוואות","מקוואות גברים ונשים")
r+=1; ir(r,"אירועים_ושיעורים","אירועים חד-פעמיים והכרזות")
r+=1; ir(r,"שיעורים_קבועים","שיעורים שבועיים (מקושרים לבית כנסת)")


# ── Synagogues (no prayer time columns, no originalNusach) ───────────────────
SYN = [
    ("id *",            "מזהה *",          "",           13),
    ("name *",          "שם *",             "",           28),
    ("neighborhood",    "שכונה",            "",           20),
    ("address_he",      "כתובת עברית",      "",           32),
    ("address_en",      "כתובת אנגלית",     "",           32),
    ("nusach *",        "נוסח *",           "ashkenaz/…", 18),
    ("phone",           "טלפון",            "",           16),
    ("rabbiName",       "שם הרב",           "",           22),
    ("rabbiPhone",      "טלפון הרב",        "",           16),
    ("gabbaiName",      "שם גבאי",          "",           22),
    ("gabbaiPhone",     "טלפון גבאי",       "",           16),
    ("latitude",        "קו רוחב",          "31.9234",    13),
    ("longitude",       "קו אורך",          "35.0123",    13),
    ("wazeLink",        "Waze",             "",           26),
    ("navigationNote",  "הערת ניווט",       "",           24),
    ("notes",           "הערות",            "",           24),
]
SYN_EX = [
    "syn-001","בית כנסת אוהל יצחק","מרכז","רחוב הרצל 12","12 Herzl St",
    "ashkenaz","052-1234567","הרב ישראל כהן","054-9876543",
    "שמעון לוי","050-1112233","31.9234","35.0123",
    "https://waze.com/ul?ll=31.9234,35.0123","","חניה בסמוך",
]
add_sheet("בתי_כנסת", SYN, SYN_EX)


# ── Prayer Times (NEW) ────────────────────────────────────────────────────────
PRAY = [
    ("synagogueId *",  "מזהה ביכ\"נ *",    "חייב להתאים ל-id בגיליון בתי_כנסת",  22),
    ("prayerType *",   "תפילה *",          "shacharit / mincha / maariv",          18),
    ("scheduleType *", "סוג לוח *",        "weekday / shabbat / friday_mincha",    22),
    ("days",           "ימים",             "1,2,3 (ריק ל-shabbat/friday_mincha)", 20),
    ("time",           "שעה קבועה",        "HH:MM — ריק אם משתמשים בעוגן",        16),
    ("anchor",         "עוגן",             "netz / shkia / chatzot / …",           24),
    ("offsetMin",      "קיזוז (דקות)",     "+20 = אחרי  |  -20 = לפני",           18),
    ("notes",          "הערות",            "",                                      28),
]
PRAY_EX_1 = ["syn-001","shacharit","weekday","1,2,3,4,5","06:30","","",""]
PRAY_EX_2 = ["syn-001","mincha","weekday","1,2,3,4,5","","shkia","-20","20 דק' לפני שקיעה"]
PRAY_EX_3 = ["syn-001","maariv","shabbat","","22:00","","","מוצאי שבת"]

ws_pray = wb.create_sheet("תפילות_מפורטות")
ws_pray.row_dimensions[1].height = 38
for ci, (en, he, _note, w) in enumerate(PRAY, 1):
    c = ws_pray.cell(row=1, column=ci, value=f"{he}\n{en}")
    hdr(c)
    ws_pray.column_dimensions[get_column_letter(ci)].width = w
for ri, row_data in enumerate([PRAY_EX_1, PRAY_EX_2, PRAY_EX_3], 2):
    ws_pray.row_dimensions[ri].height = 18
    for ci, val in enumerate(row_data, 1):
        c = ws_pray.cell(row=ri, column=ci, value=val)
        ex(c)
ws_pray.freeze_panes = "A2"


# ── Restaurants ───────────────────────────────────────────────────────────────
REST = [
    ("id *",             "מזהה *",         "",                   13),
    ("name *",           "שם *",            "",                   28),
    ("category *",       "קטגוריה *",      "meat/dairy/…",       16),
    ("neighborhood",     "שכונה",           "",                   20),
    ("address *",        "כתובת *",        "",                   32),
    ("phone",            "טלפון",           "",                   16),
    ("website",          "אתר",            "",                   26),
    ("latitude",         "קו רוחב",        "",                   13),
    ("longitude",        "קו אורך",        "",                   13),
    ("imageUrl",         "קישור תמונה",    "",                   28),
    ("activeAlert",      "התראה פעילה",    "",                   26),
    ("hours_sunday",     "ראשון",          "HH:MM-HH:MM / סגור", 16),
    ("hours_monday",     "שני",            "",                   16),
    ("hours_tuesday",    "שלישי",          "",                   16),
    ("hours_wednesday",  "רביעי",          "",                   16),
    ("hours_thursday",   "חמישי",          "",                   16),
    ("hours_friday",     "שישי",           "",                   16),
    ("hours_saturday",   "שבת",           "",                   16),
    ("kosher_issuedBy",  "גוף מכשיר",     "",                   26),
    ("kosher_certNumber","מספר תעודה",    "",                   18),
    ("kosher_level",     "רמת כשרות",     "mehadrin,glatt,…",   30),
    ("kosher_validFrom", "תוקף מ",        "YYYY-MM-DD",         16),
    ("kosher_validUntil","תוקף עד",       "YYYY-MM-DD",         16),
    ("kosher_notes",     "הערות כשרות",   "",                   28),
]
REST_EX = [
    "rest-001","מסעדת הגריל הכשר","meat","מרכז","שדרות בן גוריון 5",
    "02-1234567","https://grill.co.il","31.9210","35.0100","","",
    "09:00-23:00","09:00-23:00","09:00-23:00","09:00-23:00","09:00-23:00","09:00-14:00","סגור",
    "רבנות מקומית","2025-KSH-0042","mehadrin,bishul_israel","2025-01-01","2025-12-31","",
]
add_sheet("מסעדות_כשרות", REST, REST_EX)

# ── Mikveh ────────────────────────────────────────────────────────────────────
MIK = [
    ("id *",                "מזהה *",          "",                   13),
    ("name *",              "שם *",             "",                   28),
    ("type *",              "סוג *",           "women/men/both",     14),
    ("neighborhood",        "שכונה",            "",                   20),
    ("address *",           "כתובת *",         "",                   32),
    ("phone",               "טלפון",            "",                   16),
    ("requiresAppointment", "דרושה הזמנה",     "yes / no",           14),
    ("appointmentPhone",    "טלפון הזמנות",    "",                   18),
    ("latitude",            "קו רוחב",         "",                   13),
    ("longitude",           "קו אורך",         "",                   13),
    ("notes",               "הערות",            "",                   28),
    ("hours_sunday",        "ראשון",           "HH:MM-HH:MM / סגור", 16),
    ("hours_monday",        "שני",             "",                   16),
    ("hours_tuesday",       "שלישי",           "",                   16),
    ("hours_wednesday",     "רביעי",           "",                   16),
    ("hours_thursday",      "חמישי",           "",                   16),
    ("hours_friday",        "שישי",            "",                   16),
    ("hours_saturday",      "שבת",            "",                   16),
]
MIK_EX = [
    "mik-001","מקווה נשים מרכזי","women","מרכז","רחוב המקווה 3",
    "02-9876543","yes","02-9876543","31.9220","35.0110","כניסה מהצד הצפוני",
    "סגור","17:00-22:00","17:00-22:00","17:00-22:00","17:00-22:00","15:00-18:00","סגור",
]
add_sheet("מקוואות", MIK, MIK_EX)

# ── Events ────────────────────────────────────────────────────────────────────
EVT = [
    ("id *",        "מזהה *",        "",                  13),
    ("title *",     "כותרת *",       "",                  34),
    ("description", "תיאור",         "",                  42),
    ("category",    "קטגוריה",       "shiur/community/…", 18),
    ("startDate *", "תאריך התחלה *", "YYYY-MM-DD HH:MM",  20),
    ("endDate",     "תאריך סיום",   "YYYY-MM-DD HH:MM",  20),
    ("location",    "מיקום",         "",                  28),
    ("organizer",   "מארגן",         "",                  22),
    ("isAlert",     "התראה דחופה",  "yes / no",          12),
    ("imageUrl",    "קישור תמונה",  "",                  30),
]
EVT_EX = [
    "evt-001","שיעור גמרא שבועי","שיעור בדף היומי עם הרב כהן",
    "shiur","2025-09-15 20:00","2025-09-15 21:00",
    "בית כנסת אוהל יצחק","הרב ישראל כהן","no","",
]
add_sheet("אירועים_ושיעורים", EVT, EVT_EX)

# ── Shiurim ───────────────────────────────────────────────────────────────────
SHI = [
    ("synagogueId *", "מזהה ביכ\"נ *",  "חייב להתאים לגיליון בתי_כנסת", 22),
    ("shiurId *",     "מזהה שיעור *",   "e.g. shi-001",                  16),
    ("title *",       "שם השיעור *",    "",                              34),
    ("rabbi",         "מרצה / רב",      "",                              26),
    ("days",          "ימים",           "1,2,5  או  daily",              18),
    ("time",          "שעה",            "HH:MM",                         13),
    ("description",   "תיאור",          "",                              40),
]
SHI_EX = [
    "syn-001","shi-001","דף היומי","הרב ישראל כהן","1,2,3,4,5","20:00","שיעור גמרא יומי, 45 דקות",
]
add_sheet("שיעורים_קבועים", SHI, SHI_EX)

# ── Move instructions to front ────────────────────────────────────────────────
wb.move_sheet("הוראות", offset=-(len(wb.sheetnames) - 1))

out = r"C:\Temp\kehila_base.xlsx"
wb.save(out)
print("Saved:", out)
