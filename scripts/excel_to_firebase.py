"""
excel_to_firebase.py
--------------------
Converts kehila_data_template.xlsm (filled by city contact) into
Firebase-ready JSON files that can be uploaded via the Admin SDK.

Usage:
    python scripts/excel_to_firebase.py <path-to-xlsx-or-xlsm> [--city <cityId>]

Outputs (written to ./firebase_import/):
    synagogues.json
    restaurants.json
    mikveh.json
    events.json

Then upload with:
    python scripts/upload_to_firebase.py
"""

import sys
import json
import re
import argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("Run:  pip install openpyxl")

# ─── helpers ────────────────────────────────────────────────────────────────

def cell_str(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def opening_hours(row_dict, days=("sunday","monday","tuesday","wednesday","thursday","friday","saturday")):
    h = {}
    for d in days:
        v = row_dict.get(f"hours_{d}", "").strip()
        if v:
            h[d] = v
    return h


def sheet_to_dicts(ws):
    """Read a sheet; row 1 = header (Hebrew\nen_key), row 2 = example (skip), rows 3+ = data."""
    headers = []
    for cell in ws[1]:
        v = cell.value or ""
        # header format: "Hebrew label\nen_key" — take the second line (en key)
        parts = str(v).split("\n")
        key = (parts[1] if len(parts) > 1 else parts[0]).strip().rstrip("*").strip()
        headers.append(key)

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=False):
        if all(c.value is None for c in row):
            continue
        d = {headers[i]: cell_str(row[i]) for i in range(min(len(headers), len(row)))}
        rows.append(d)
    return rows


# ─── prayer times ────────────────────────────────────────────────────────────

def _parse_days_list(raw: str) -> list[int]:
    """'1,2,5' or '1-5' → [1, 2, 5] / [1, 2, 3, 4, 5]. Returns [] for empty/invalid."""
    raw = raw.strip()
    if not raw or raw.lower() == "daily":
        return list(range(1, 8))
    nums = [int(x) for x in re.findall(r"\d+", raw) if 1 <= int(x) <= 7]
    return nums


def _build_prayer_slot(d: dict) -> dict:
    """Convert one row from תפילות_מפורטות into a PrayerTimeSlot dict."""
    slot: dict = {}
    if d.get("time"):
        slot["time"] = d["time"]
    if d.get("anchor"):
        slot["anchor"] = d["anchor"]
    if d.get("offsetMin"):
        try:
            slot["offsetMin"] = int(d["offsetMin"])
        except ValueError:
            pass

    schedule_type = d.get("scheduleType", "weekday")
    if schedule_type == "shabbat":
        slot["days"] = [7]
    elif schedule_type == "friday_mincha":
        slot["days"] = [6]
    else:
        days = _parse_days_list(d.get("days", ""))
        if days:
            slot["days"] = days

    if d.get("notes"):
        slot["notes"] = d["notes"]
    return slot


def build_schedules(prayer_rows: list[dict]) -> tuple[dict, dict]:
    """
    Build (weeklySchedule, shabbatSchedule) dicts from raw prayer sheet rows.
    weeklySchedule  keys: shacharit | mincha | maariv
    shabbatSchedule keys: shacharit | mincha | maariv | minchaFriday
    """
    weekly: dict[str, list] = {}
    shabbat: dict[str, list] = {}

    for row in prayer_rows:
        prayer_type   = row.get("prayerType", "").strip()
        schedule_type = row.get("scheduleType", "weekday").strip()
        if not prayer_type:
            continue

        slot = _build_prayer_slot(row)

        if schedule_type == "weekday":
            weekly.setdefault(prayer_type, []).append(slot)
        elif schedule_type == "shabbat":
            shabbat.setdefault(prayer_type, []).append(slot)
        elif schedule_type == "friday_mincha":
            shabbat.setdefault("minchaFriday", []).append(slot)

    return weekly, shabbat


# ─── converters ─────────────────────────────────────────────────────────────

def convert_synagogue(
    d: dict,
    cityId: str,
    prayers_by_syn: dict,
    shiurim_by_syn: dict,
) -> dict | None:
    if not d.get("id") or not d.get("name"):
        return None

    syn_id = d["id"]
    weekly, shabbat = prayers_by_syn.get(syn_id, ({}, {}))

    address = {}
    if d.get("address_he"):
        address["he"] = d["address_he"]
    if d.get("address_en"):
        address["en"] = d["address_en"]

    syn = {
        "id":             syn_id,
        "cityId":         cityId,
        "name":           d["name"],
        "nusach":         d.get("nusach", "other"),
        "address":        address,
        "weeklySchedule": weekly,
    }

    for src, dst in [
        ("neighborhood",   "neighborhood"),
        ("phone",          "phone"),
        ("rabbiName",      "rabbiName"),
        ("rabbiPhone",     "rabbiPhone"),
        ("gabbaiName",     "gabbaiName"),
        ("gabbaiPhone",    "gabbaiPhone"),
        ("wazeLink",       "wazeLink"),
        ("navigationNote", "navigationNote"),
        ("notes",          "notes"),
    ]:
        if d.get(src):
            syn[dst] = d[src]

    for coord in ("latitude", "longitude"):
        if d.get(coord):
            try:
                syn[coord] = float(d[coord])
            except ValueError:
                pass

    if shabbat:
        syn["shabbatSchedule"] = shabbat

    shiurim = shiurim_by_syn.get(syn_id, [])
    if shiurim:
        syn["shiurim"] = shiurim

    return syn


def convert_restaurant(d: dict, cityId: str) -> dict | None:
    if not d.get("id") or not d.get("name"):
        return None

    cert = {}
    if d.get("kosher_issuedBy"):
        levels = [l.strip() for l in d.get("kosher_level", "").split(",") if l.strip()]
        cert = {
            "id":          d["id"] + "-cert",
            "issuedBy":    d["kosher_issuedBy"],
            "kosherLevel": levels or ["regular"],
            "validFrom":   d.get("kosher_validFrom", ""),
            "validUntil":  d.get("kosher_validUntil", ""),
            "isActive":    True,
        }
        if d.get("kosher_certNumber"):
            cert["certNumber"] = d["kosher_certNumber"]
        if d.get("kosher_notes"):
            cert["notes"] = d["kosher_notes"]

    rest = {
        "id":                 d["id"],
        "cityId":             cityId,
        "name":               d["name"],
        "category":           d.get("category", ""),
        "address":            d.get("address", ""),
        "openingHours":       opening_hours(d),
        "kosherCertificates": [cert] if cert else [],
    }

    for src, dst in [
        ("neighborhood", "neighborhood"),
        ("phone",        "phone"),
        ("website",      "website"),
        ("imageUrl",     "imageUrl"),
        ("activeAlert",  "activeAlert"),
    ]:
        if d.get(src):
            rest[dst] = d[src]

    for coord in ("latitude", "longitude"):
        if d.get(coord):
            try:
                rest[coord] = float(d[coord])
            except ValueError:
                pass

    return rest


def convert_mikveh(d: dict, cityId: str) -> dict | None:
    if not d.get("id") or not d.get("name"):
        return None

    mik = {
        "id":                  d["id"],
        "cityId":              cityId,
        "name":                d["name"],
        "type":                d.get("type", "both"),
        "address":             d.get("address", ""),
        "openingHours":        opening_hours(d),
        "requiresAppointment": d.get("requiresAppointment", "no").lower() == "yes",
    }

    for src, dst in [
        ("neighborhood",     "neighborhood"),
        ("phone",            "phone"),
        ("appointmentPhone", "appointmentPhone"),
        ("notes",            "notes"),
    ]:
        if d.get(src):
            mik[dst] = d[src]

    for coord in ("latitude", "longitude"):
        if d.get(coord):
            try:
                mik[coord] = float(d[coord])
            except ValueError:
                pass

    return mik


def convert_event(d: dict, cityId: str) -> dict | None:
    if not d.get("id") or not d.get("title"):
        return None

    evt = {
        "id":          d["id"],
        "cityId":      cityId,
        "title":       d["title"],
        "description": d.get("description", ""),
        "category":    d.get("category", "community"),
        "startDate":   d.get("startDate", ""),
        "isAlert":     d.get("isAlert", "no").lower() == "yes",
        "createdBy":   "admin",
        "createdAt":   datetime.utcnow().isoformat() + "Z",
    }

    for src, dst in [
        ("endDate",   "endDate"),
        ("location",  "location"),
        ("organizer", "organizer"),
        ("imageUrl",  "imageUrl"),
    ]:
        if d.get(src):
            evt[dst] = d[src]

    return evt


def convert_shiur(d: dict) -> dict | None:
    if not d.get("shiurId") or not d.get("title"):
        return None

    days_raw = d.get("days", "").strip()
    if days_raw.lower() == "daily":
        days: list | str = "daily"
    else:
        days = [int(x) for x in re.findall(r"\d", days_raw) if x]

    return {
        "id":          d["shiurId"],
        "title":       d["title"],
        "rabbi":       d.get("rabbi", ""),
        "days":        days,
        "time":        d.get("time", ""),
        "description": d.get("description", "") or None,
    }


# ─── main ────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Convert kehila Excel template to Firebase JSON")
    ap.add_argument("xlsx", help="Path to filled kehila_data_template.xlsm (or .xlsx)")
    ap.add_argument("--city", default="city-1", help="Firebase cityId (default: city-1)")
    ap.add_argument("--out",  default="firebase_import", help="Output directory")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, keep_vba=True)
    city = args.city
    out_dir = Path(args.out)
    out_dir.mkdir(exist_ok=True)

    # ── Prayer times (needed before synagogues) ──────────────────────────────
    prayers_raw: dict[str, list] = {}
    if "תפילות_מפורטות" in wb.sheetnames:
        for d in sheet_to_dicts(wb["תפילות_מפורטות"]):
            syn_id = d.get("synagogueId", "").strip()
            if syn_id:
                prayers_raw.setdefault(syn_id, []).append(d)

    prayers_by_syn: dict[str, tuple] = {
        syn_id: build_schedules(rows)
        for syn_id, rows in prayers_raw.items()
    }

    # ── Shiurim (needed before synagogues) ───────────────────────────────────
    shiurim_by_syn: dict[str, list] = {}
    if "שיעורים_קבועים" in wb.sheetnames:
        for d in sheet_to_dicts(wb["שיעורים_קבועים"]):
            syn_id = d.get("synagogueId", "").strip()
            if not syn_id:
                continue
            shi = convert_shiur(d)
            if shi:
                shiurim_by_syn.setdefault(syn_id, []).append(shi)

    # ── Synagogues ───────────────────────────────────────────────────────────
    syns = []
    if "בתי_כנסת" in wb.sheetnames:
        for d in sheet_to_dicts(wb["בתי_כנסת"]):
            obj = convert_synagogue(d, city, prayers_by_syn, shiurim_by_syn)
            if obj:
                syns.append(obj)

    # ── Restaurants ──────────────────────────────────────────────────────────
    rests = []
    if "מסעדות_כשרות" in wb.sheetnames:
        for d in sheet_to_dicts(wb["מסעדות_כשרות"]):
            obj = convert_restaurant(d, city)
            if obj:
                rests.append(obj)

    # ── Mikveh ───────────────────────────────────────────────────────────────
    mikveh = []
    if "מקוואות" in wb.sheetnames:
        for d in sheet_to_dicts(wb["מקוואות"]):
            obj = convert_mikveh(d, city)
            if obj:
                mikveh.append(obj)

    # ── Events ───────────────────────────────────────────────────────────────
    events = []
    if "אירועים_ושיעורים" in wb.sheetnames:
        for d in sheet_to_dicts(wb["אירועים_ושיעורים"]):
            obj = convert_event(d, city)
            if obj:
                events.append(obj)

    # ── Write JSON ───────────────────────────────────────────────────────────
    files = {
        "synagogues.json":  syns,
        "restaurants.json": rests,
        "mikveh.json":      mikveh,
        "events.json":      events,
    }
    for fname, data in files.items():
        p = out_dir / fname
        p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"  {fname}: {len(data)} records")

    print(f"\nDone. Files written to ./{args.out}/")
    print("\nTo upload to Firebase run:")
    print("  python scripts/upload_to_firebase.py")


if __name__ == "__main__":
    main()
